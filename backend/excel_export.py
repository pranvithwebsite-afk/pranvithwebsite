"""
This module provides functionality to export data to Excel files with professional formatting.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
import io

def create_excel_report(title: str, report_data: list, summary_data: dict = None, sheet_name: str = "Report"):
    """
    Creates an Excel report from a list of dictionaries.

    Args:
        title (str): The title of the report.
        report_data (list): A list of dictionaries, where each dictionary represents a row.
        summary_data (dict): A dictionary with summary data.
        sheet_name (str): The name of the worksheet.

    Returns:
        bytes: The Excel file in memory.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Add report title and generation time
    ws.merge_cells('A1:F1')
    ws['A1'].value = title
    ws['A1'].font = Font(size=18, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A2:F2')
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws['A2'].value = f"Generated on: {now}"
    ws['A2'].font = Font(size=10, italic=True)
    ws['A2'].alignment = Alignment(horizontal='center')

    # Add summary section
    if summary_data:
        summary_start_row = 4
        ws.merge_cells(f'A{summary_start_row}:B{summary_start_row}')
        ws[f'A{summary_start_row}'].value = "Summary"
        ws[f'A{summary_start_row}'].font = Font(size=14, bold=True)
        summary_start_row += 1
        for key, value in summary_data.items():
            ws.cell(row=summary_start_row, column=1, value=key).font = Font(bold=True)
            ws.cell(row=summary_start_row, column=2, value=value)
            summary_start_row += 1

    data_start_row = ws.max_row + 2 if summary_data else 4

    if not report_data:
        # Handle case with no data
        ws.merge_cells(f'A{data_start_row}:F{data_start_row}')
        ws[f'A{data_start_row}'].value = "No data available for the selected criteria."
        ws[f'A{data_start_row}'].font = Font(size=12)
        ws[f'A{data_start_row}'].alignment = Alignment(horizontal='center')
    else:
        # Headers
        headers = list(report_data[0].keys())
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F4F4F", end_color="4F4F4F", fill_type="solid")

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=data_start_row, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Freeze header row
        ws.freeze_panes = ws.cell(row=data_start_row + 1, column=1)

        # Data rows
        alt_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        for row_idx, row_data in enumerate(report_data, start=data_start_row + 1):
            for col_idx, value in enumerate(row_data.values(), start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)
            
            # Apply alternating row color
            if (row_idx - data_start_row) % 2 != 0:
                for cell in ws[row_idx]:
                    cell.fill = alt_fill
        
        # Auto-adjust column widths
        for col_idx, header in enumerate(headers, 1):
            column_letter = get_column_letter(col_idx)
            max_length = len(header)
            # Check all rows for max length, including the header
            for row_num in range(data_start_row, ws.max_row + 1):
                cell_value = ws.cell(row=row_num, column=col_idx).value
                if cell_value is not None:
                    max_length = max(max_length, len(str(cell_value)))
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width

    # Save to a memory buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
